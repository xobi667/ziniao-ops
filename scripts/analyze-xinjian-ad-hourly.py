#!/usr/bin/env python3
"""Analyze Xinjian ERP hourly ad exports without browser session access."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable


METRIC_SYNONYMS = {
    "store": ["店铺", "店铺名称", "shopname", "shop_name", "store", "storename", "alias"],
    "hour": ["小时", "时段", "时间", "时间段", "小时段", "分时", "hour", "reportdatestr", "reporthour"],
    "date": ["日期", "统计日期", "date", "reportdate", "reporthour", "时间"],
    "cost": ["花费", "广告花费", "消耗", "adcost", "cost", "spend"],
    "revenue": ["广告销售额", "广告店铺收入", "销售额", "营业额", "adgmv", "gmv", "revenue", "sales"],
    "orders": ["广告订单", "广告出单数", "订单", "出单", "broadorder", "directorder", "order"],
    "clicks": ["点击数", "点击", "adclick", "click", "clicks"],
    "impressions": ["展示次数", "展示", "展现", "曝光", "impression", "impressions", "views"],
    "roas": ["roas", "roi", "投入产出"],
    "ctr": ["ctr", "点击率"],
    "cr": ["cr", "转化率"],
    "cpc": ["cpc"],
}

NUMERIC_KEYS = ["cost", "revenue", "orders", "clicks", "impressions"]
EXCLUDE_DIRS = {".git", ".upstreams", ".ziniao-ops", "node_modules", "__pycache__", "reports.local"}
EXTS = {".csv", ".json", ".xlsx"}


def classify_source(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".json":
        try:
            value = json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception:  # noqa: BLE001 - source classification must not block parsing later.
            return "json_file"
        if isinstance(value, dict) and "code" in value and "msg" in value:
            return "xinjian_endpoint_response_json"
        return "json_export"
    if suffix == ".xlsx":
        return "xlsx_export"
    if suffix == ".csv":
        return "csv_export"
    return "unknown"


def norm(value: Any) -> str:
    return re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff]+", "", str(value or "")).lower()


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "--", "null", "None"}:
        return None
    text = text.replace(",", "")
    text = re.sub(r"[^\d.\-]+", "", text)
    if not text or text in {"-", "."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if number > 10_000_000_000:
            return dt.datetime.fromtimestamp(number / 1000).date()
        if number > 10_000_000:
            return dt.datetime.fromtimestamp(number).date()
        if 20_000 < number < 80_000:
            origin = dt.datetime(1899, 12, 30)
            return (origin + dt.timedelta(days=number)).date()
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("/", "-").replace(".", "-")
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        try:
            return dt.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


def parse_hour(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.hour
    if isinstance(value, dt.time):
        return value.hour
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if 0 < number < 1:
            return int(number * 24)
        if 0 <= number <= 23:
            return int(number)
        if number > 10_000_000_000:
            return dt.datetime.fromtimestamp(number / 1000).hour
        if number > 10_000_000:
            return dt.datetime.fromtimestamp(number).hour
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", text):
        return None
    patterns = [
        r"\b([01]?\d|2[0-3])\s*[:：点时]",
        r"^([01]?\d|2[0-3])\s*(?:-|~|至|到)",
        r"\b([01]?\d|2[0-3])\s*-\s*(?:[01]?\d|2[0-3])",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return int(match.group(1))
    parsed_date = parse_date(text)
    if parsed_date and re.search(r"\d{1,2}:\d{2}", text):
        match = re.search(r"\b([01]?\d|2[0-3]):\d{2}", text)
        if match:
            return int(match.group(1))
    return None


def read_csv(path: Path) -> list[list[Any]]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [row for row in csv.reader(handle)]
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, "unsupported encoding")


def read_xlsx(path: Path) -> list[tuple[str, list[list[Any]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for .xlsx exports") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: list[tuple[str, list[list[Any]]]] = []
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        sheets.append((ws.title, rows))
    return sheets


def flatten_json_records(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        result: list[dict[str, Any]] = []
        for item in value:
            result.extend(flatten_json_records(item))
        return result
    if isinstance(value, dict):
        if all(not isinstance(v, (list, dict)) for v in value.values()):
            return [value]
        result = []
        if "currentData" in value:
            result.extend(flatten_json_records(value["currentData"]))
            return result
        for key in ("data", "list", "rows", "currentData", "preData", "records"):
            if key in value:
                result.extend(flatten_json_records(value[key]))
        return result
    return []


def rows_from_json(path: Path) -> list[tuple[str, list[list[Any]]]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    records = flatten_json_records(data)
    if not records:
        return []
    headers = sorted({key for record in records for key in record.keys()})
    rows = [headers]
    for record in records:
        rows.append([record.get(header) for header in headers])
    return [("json", rows)]


def detect_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    best: tuple[int, int, dict[str, int]] | None = None
    for idx, row in enumerate(rows[:50]):
        mapping: dict[str, int] = {}
        normalized_cells = [norm(cell) for cell in row]
        for key, synonyms in METRIC_SYNONYMS.items():
            normalized_synonyms = [norm(s) for s in synonyms]
            for col_idx, cell in enumerate(normalized_cells):
                if not cell:
                    continue
                if any(s == cell or (len(s) >= 3 and s in cell) for s in normalized_synonyms):
                    mapping.setdefault(key, col_idx)
                    break
        score = len(mapping)
        if score >= 3 and (best is None or score > best[1]):
            best = (idx, score, mapping)
    if not best:
        return None
    return best[0], best[2]


def iter_candidate_files(roots: Iterable[Path]) -> Iterable[Path]:
    for root in roots:
        if not root.exists():
            continue
        if root.is_file():
            if root.suffix.lower() in EXTS:
                yield root
            continue
        for current, dirs, files in os.walk(root):
            dirs[:] = [d for d in dirs if d not in EXCLUDE_DIRS]
            for name in files:
                path = Path(current) / name
                if path.suffix.lower() not in EXTS:
                    continue
                if re.search(r"xinjian|心舰|广告|hour|分时|blue|mrs|cat", name, re.I):
                    yield path


def store_matches(store_value: str, requested: list[str]) -> bool:
    if not requested:
        return True
    current = store_value.lower()
    if not current:
        return False
    for item in requested:
        target = item.lower()
        if target in current or current == target:
            return True
    return False


def load_tables(path: Path) -> list[tuple[str, list[list[Any]]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return [("csv", read_csv(path))]
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix == ".json":
        return rows_from_json(path)
    raise RuntimeError(f"unsupported file type {suffix}; use .xlsx, .csv, or .json")


def analyze_file(path: Path, stores: list[str], start: dt.date | None, end: dt.date | None) -> tuple[list[dict[str, Any]], list[str]]:
    records: list[dict[str, Any]] = []
    notes: list[str] = []
    try:
        tables = load_tables(path)
    except Exception as exc:  # noqa: BLE001 - report and continue across files.
        return [], [f"{path}: {exc}"]
    for sheet_name, rows in tables:
        header = detect_header(rows)
        if not header:
            continue
        header_idx, mapping = header
        for row in rows[header_idx + 1 :]:
            if not any(cell not in (None, "") for cell in row):
                continue
            def cell(key: str) -> Any:
                idx = mapping.get(key)
                return row[idx] if idx is not None and idx < len(row) else None

            has_store_column = "store" in mapping
            store = str(cell("store") or "").strip()
            if stores and has_store_column and not store_matches(store, stores):
                continue
            if stores and not has_store_column:
                if len(stores) == 1:
                    store = stores[0]
                else:
                    continue
            hour = parse_hour(cell("hour")) if "hour" in mapping else None
            if hour is None:
                hour = parse_hour(cell("date"))
            if hour is None:
                continue
            row_date = parse_date(cell("date")) or parse_date(cell("hour"))
            if row_date and start and row_date < start:
                continue
            if row_date and end and row_date > end:
                continue
            record = {
                "file": str(path),
                "sheet": sheet_name,
                "source_type": classify_source(path),
                "store": store or (stores[0] if len(stores) == 1 else "UNKNOWN_STORE"),
                "hour": hour,
                "date": row_date.isoformat() if row_date else "",
            }
            for key in NUMERIC_KEYS:
                record[key] = to_float(cell(key)) or 0.0
            records.append(record)
    return records, notes


def aggregate(records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[tuple[str, int], dict[str, Any]] = {}
    for record in records:
        key = (record["store"], record["hour"])
        item = grouped.setdefault(
            key,
            {
                "store": record["store"],
                "hour": record["hour"],
                "rows": 0,
                "dates": set(),
                "cost": 0.0,
                "revenue": 0.0,
                "orders": 0.0,
                "clicks": 0.0,
                "impressions": 0.0,
            },
        )
        item["rows"] += 1
        if record.get("date"):
            item["dates"].add(record["date"])
        for metric in NUMERIC_KEYS:
            item[metric] += float(record.get(metric) or 0)
    by_store: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in grouped.values():
        cost = item["cost"]
        revenue = item["revenue"]
        clicks = item["clicks"]
        impressions = item["impressions"]
        orders = item["orders"]
        item["roas"] = revenue / cost if cost > 0 else 0.0
        item["ctr"] = clicks / impressions if impressions > 0 else 0.0
        item["cr"] = orders / clicks if clicks > 0 else 0.0
        item["cpc"] = cost / clicks if clicks > 0 else 0.0
        item["date_count"] = len(item["dates"])
        item["dates"] = sorted(item["dates"])
        by_store[item["store"]].append(item)
    for store_items in by_store.values():
        store_items.sort(key=lambda x: x["hour"])
    return by_store


def best_rows(by_store: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    result = []
    for store, items in by_store.items():
        activity_items = [
            i
            for i in items
            if any(float(i.get(metric) or 0) > 0 for metric in NUMERIC_KEYS)
        ]
        if not activity_items:
            dates = sorted({date for item in items for date in item.get("dates", [])})
            item = {
                "store": store,
                "hour": None,
                "rows": sum(int(i.get("rows") or 0) for i in items),
                "dates": dates,
                "cost": sum(float(i.get("cost") or 0) for i in items),
                "revenue": sum(float(i.get("revenue") or 0) for i in items),
                "orders": sum(float(i.get("orders") or 0) for i in items),
                "clicks": sum(float(i.get("clicks") or 0) for i in items),
                "impressions": sum(float(i.get("impressions") or 0) for i in items),
                "roas": 0.0,
                "ctr": 0.0,
                "cr": 0.0,
                "cpc": 0.0,
                "date_count": len(dates),
                "no_activity": True,
                "reason": "全部时段无花费、销售额、订单、点击和展示，无法判断最佳投放时段",
            }
            result.append(item)
            continue
        candidates = [i for i in items if i["cost"] > 0 and (i["revenue"] > 0 or i["orders"] > 0 or i["clicks"] > 0)]
        if not candidates:
            candidates = activity_items
        best = max(
            candidates,
            key=lambda i: (
                i["roas"],
                i["orders"],
                i["revenue"],
                i["clicks"],
                -i["cpc"],
            ),
        )
        item = dict(best)
        item["no_activity"] = False
        item["reason"] = "ROAS最高，订单/销售额作为并列判断" if best["cost"] > 0 else "缺少花费，按可用互动/订单判断"
        result.append(item)
    result.sort(key=lambda x: x["store"].lower())
    return result


def fmt_money(value: float) -> str:
    return f"{value:,.2f}"


def fmt_int(value: float) -> str:
    return f"{value:,.0f}"


def fmt_pct(value: float) -> str:
    return f"{value * 100:.2f}%"


def hour_label(hour: Any) -> str:
    if hour is None or hour == "":
        return "无有效时段"
    return f"{hour:02d}:00-{hour:02d}:59"


def markdown_report(result: dict[str, Any]) -> str:
    lines = [
        "# 心舰 ERP 产品广告分时分析",
        "",
        f"- 数据文件数: {len(result['files_used'])}",
        f"- 数据来源类型: {', '.join(result.get('source_types') or []) or '-'}",
        f"- 记录数: {result['record_count']}",
        f"- 日期范围: {result.get('start_date') or '-'} ~ {result.get('end_date') or '-'}",
        "",
        "## 最佳时段",
        "",
        "| 店铺 | 最佳时段 | ROAS | 广告销售额 | 广告花费 | 广告订单 | 点击 | 展示 | CTR | CR | CPC | 样本行数 | 判断 |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---|",
    ]
    for row in result["best"]:
        lines.append(
            "| {store} | {hour} | {roas:.2f} | {revenue} | {cost} | {orders} | {clicks} | {impressions} | {ctr} | {cr} | {cpc} | {rows} | {reason} |".format(
                store=row["store"],
                hour=hour_label(row["hour"]),
                roas=row["roas"],
                revenue=fmt_money(row["revenue"]),
                cost=fmt_money(row["cost"]),
                orders=fmt_int(row["orders"]),
                clicks=fmt_int(row["clicks"]),
                impressions=fmt_int(row["impressions"]),
                ctr=fmt_pct(row["ctr"]),
                cr=fmt_pct(row["cr"]),
                cpc=fmt_money(row["cpc"]),
                rows=row["rows"],
                reason=row["reason"],
            )
        )
    lines += ["", "## 分时明细", ""]
    for store, items in result["hourly"].items():
        lines += [
            f"### {store}",
            "",
            "| 时段 | ROAS | 广告销售额 | 广告花费 | 广告订单 | 点击 | 展示 | CTR | CR | CPC | 样本行数 |",
            "|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
        ]
        for row in items:
            lines.append(
                "| {hour} | {roas:.2f} | {revenue} | {cost} | {orders} | {clicks} | {impressions} | {ctr} | {cr} | {cpc} | {rows} |".format(
                    hour=hour_label(row["hour"]),
                    roas=row["roas"],
                    revenue=fmt_money(row["revenue"]),
                    cost=fmt_money(row["cost"]),
                    orders=fmt_int(row["orders"]),
                    clicks=fmt_int(row["clicks"]),
                    impressions=fmt_int(row["impressions"]),
                    ctr=fmt_pct(row["ctr"]),
                    cr=fmt_pct(row["cr"]),
                    cpc=fmt_money(row["cpc"]),
                    rows=row["rows"],
                )
            )
        lines.append("")
    if result["notes"]:
        lines += ["## 备注", ""]
        lines.extend(f"- {note}" for note in result["notes"])
        lines.append("")
    return "\n".join(lines)


def write_xlsx_report(result: dict[str, Any], output: Path, records: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for .xlsx output") from exc

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    best_fill = PatternFill("solid", fgColor="E2F0D9")

    def style_sheet(ws) -> None:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 40))
                cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    def add_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
        ws.append(headers)
        for row in rows:
            ws.append(row)
        style_sheet(ws)

    ws_best = wb.active
    ws_best.title = "最佳时段"
    best_headers = [
        "店铺",
        "最佳时段",
        "ROAS",
        "广告销售额",
        "广告花费",
        "广告订单",
        "点击",
        "展示",
        "CTR",
        "CR",
        "CPC",
        "样本行数",
        "样本日期数",
        "判断",
    ]
    best_data = []
    for row in result["best"]:
        best_data.append(
            [
                row["store"],
                hour_label(row["hour"]),
                round(row["roas"], 4),
                round(row["revenue"], 2),
                round(row["cost"], 2),
                round(row["orders"], 0),
                round(row["clicks"], 0),
                round(row["impressions"], 0),
                round(row["ctr"], 4),
                round(row["cr"], 4),
                round(row["cpc"], 4),
                row["rows"],
                row["date_count"],
                row["reason"],
            ]
        )
    add_rows(ws_best, best_headers, best_data)
    for row in ws_best.iter_rows(min_row=2):
        for cell in row:
            cell.fill = best_fill

    ws_hourly = wb.create_sheet("分时明细")
    hourly_headers = [
        "店铺",
        "时段",
        "ROAS",
        "广告销售额",
        "广告花费",
        "广告订单",
        "点击",
        "展示",
        "CTR",
        "CR",
        "CPC",
        "样本行数",
        "样本日期数",
    ]
    hourly_data = []
    for store, items in result["hourly"].items():
        for row in items:
            hourly_data.append(
                [
                    store,
                    hour_label(row["hour"]),
                    round(row["roas"], 4),
                    round(row["revenue"], 2),
                    round(row["cost"], 2),
                    round(row["orders"], 0),
                    round(row["clicks"], 0),
                    round(row["impressions"], 0),
                    round(row["ctr"], 4),
                    round(row["cr"], 4),
                    round(row["cpc"], 4),
                    row["rows"],
                    row["date_count"],
                ]
            )
    add_rows(ws_hourly, hourly_headers, hourly_data)

    ws_raw = wb.create_sheet("原始记录")
    raw_headers = ["店铺", "日期", "时段", "广告销售额", "广告花费", "广告订单", "点击", "展示", "来源类型", "来源文件", "工作表"]
    raw_data = []
    for record in records:
        raw_data.append(
            [
                record["store"],
                record["date"],
                hour_label(record["hour"]),
                round(record["revenue"], 2),
                round(record["cost"], 2),
                round(record["orders"], 0),
                round(record["clicks"], 0),
                round(record["impressions"], 0),
                record.get("source_type", ""),
                record["file"],
                record["sheet"],
            ]
        )
    add_rows(ws_raw, raw_headers, raw_data)

    ws_notes = wb.create_sheet("说明")
    note_rows = [
        ["项目", "内容"],
        ["日期范围", f"{result.get('start_date') or '-'} ~ {result.get('end_date') or '-'}"],
        ["请求店铺", ", ".join(result.get("stores_requested") or [])],
        ["数据文件数", len(result.get("files_used") or [])],
        ["数据来源类型", ", ".join(result.get("source_types") or [])],
        ["记录数", result.get("record_count", 0)],
        ["最佳时段口径", "优先 ROAS；并列时看广告订单、广告销售额、点击、CPC"],
        ["日期过滤", "已放宽到文件内全部可识别小时数据" if result.get("date_filter_relaxed") else "使用指定日期范围"],
        ["数据来源", "\n".join(result.get("files_used") or [])],
        ["备注", "\n".join(result.get("notes") or [])],
    ]
    for row in note_rows:
        ws_notes.append(row)
    style_sheet(ws_notes)
    for row in ws_notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for ws in (ws_best, ws_hourly):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column == 3:
                    cell.number_format = "0.00"
                elif cell.column in (9, 10):
                    cell.number_format = "0.00%"
                elif cell.column in (4, 5, 11):
                    cell.number_format = "#,##0.00"
                elif cell.column in (6, 7, 8, 12, 13):
                    cell.number_format = "#,##0"
    for row in ws_raw.iter_rows(min_row=2):
        for cell in row:
            if cell.column in (4, 5):
                cell.number_format = "#,##0.00"
            elif cell.column in (6, 7, 8):
                cell.number_format = "#,##0"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze Xinjian ERP hourly ad exports.")
    parser.add_argument("--input", action="append", default=[], help="CSV/XLSX/JSON export path.")
    parser.add_argument("--search-root", action="append", default=[], help="Directory to scan when --input is omitted.")
    parser.add_argument("--store", action="append", default=[], help="Store name filter.")
    parser.add_argument("--start-date", default="", help="Inclusive YYYY-MM-DD start date.")
    parser.add_argument("--end-date", default="", help="Inclusive YYYY-MM-DD end date.")
    parser.add_argument("--days", type=int, default=7, help="Recent days used when dates are omitted.")
    parser.add_argument("--output", default="", help="Markdown output path.")
    parser.add_argument("--xlsx-output", default="", help="XLSX output path.")
    parser.add_argument("--json", action="store_true", help="Print JSON result.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    today = dt.date.today()
    end = dt.date.fromisoformat(args.end_date) if args.end_date else today
    start = dt.date.fromisoformat(args.start_date) if args.start_date else end - dt.timedelta(days=max(args.days, 1) - 1)

    input_paths = [Path(p) for p in args.input]
    if not input_paths:
        roots = [Path(p) for p in args.search_root] or [Path.cwd()]
        input_paths = list(dict.fromkeys(iter_candidate_files(roots)))

    all_records: list[dict[str, Any]] = []
    notes: list[str] = []
    files_used: list[str] = []
    for path in input_paths:
        records, file_notes = analyze_file(path, args.store, start, end)
        notes.extend(file_notes)
        if records:
            all_records.extend(records)
            files_used.append(str(path))

    date_filter_relaxed = False
    if not all_records and (args.start_date or args.end_date or args.days):
        for path in input_paths:
            records, _ = analyze_file(path, args.store, None, None)
            if records:
                all_records.extend(records)
                files_used.append(str(path))
        if all_records:
            date_filter_relaxed = True
            notes.append("指定日期范围内没有匹配记录，已退回使用文件内全部可识别小时数据。")

    by_store = aggregate(all_records)
    best = best_rows(by_store) if by_store else []
    result = {
        "ok": bool(best),
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "date_filter_relaxed": date_filter_relaxed,
        "stores_requested": args.store,
        "files_scanned": [str(p) for p in input_paths],
        "files_used": files_used,
        "source_types": sorted({classify_source(Path(path)) for path in files_used}),
        "record_count": len(all_records),
        "best": best,
        "hourly": by_store,
        "notes": notes,
    }

    if result["ok"] and args.output:
        output = Path(args.output)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(markdown_report(result), encoding="utf-8")
        result["output"] = str(output)
    if result["ok"] and args.xlsx_output:
        xlsx_output = Path(args.xlsx_output)
        write_xlsx_report(result, xlsx_output, all_records)
        result["xlsx_output"] = str(xlsx_output)

    if args.json:
        print(json.dumps(result, ensure_ascii=True, indent=2))
    elif result["ok"] and args.output:
        print(f"Report written: {args.output}")
    else:
        print(markdown_report(result))
    return 0 if result["ok"] else 2


if __name__ == "__main__":
    sys.exit(main())
